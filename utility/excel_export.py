from openpyxl import Workbook
import os


def save_to_xlsx(experiments, filepath):
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Average"
    get_averages(experiments, ws1)

    ws2 = wb.create_sheet(title="Efficiency")
    get_efficiencies(experiments, ws2)

    for i, experiment in enumerate(experiments.values()):
        ws = wb.create_sheet(title="Experiment %d" % i)
        get_data(experiment, ws)

    wb.save(filepath)


def get_averages(experiments, worksheet):
    col_titles = ['Experiment', 'Time (s)', 'Film Thickness (mm)', 'Film Area (cm2)', 'Isc (A)', 'dIsc (A)', 'Voc (V)',
                  'dVoc (V)', 'Pmax (W)', 'dPmax (W)', 'FF', 'dFF', 'Tavg (C)', 'dTavg (C)', 'I1avg (W/m2)',
                  'dI1avg (W/m2)', 'I2avg (W/m2)', 'dI2avg (W/m2)', 'I3avg (W/m2)', 'dI3avg (W/m2)', 'I4avg (W/m2)',
                  'dI4avg (W/m2)']
    exp_vals = ['Short Circuit Current I_sc (A)', 'Open Circuit Voltage V_oc (V)', 'Maximum Power P_max (W)',
                'Fill Factor', 'Average Temperature T_avg (C)', 'Average Irradiance I_1_avg (W/m2)',
                'Average Irradiance I_2_avg (W/m2)', 'Average Irradiance I_3_avg (W/m2)',
                'Average Irradiance I_4_avg (W/m2)']
    for col, title in enumerate(col_titles, 1):
        worksheet.cell(row=1, column=col).value = title
    for row, experiment in enumerate(experiments.values(), 2):
        worksheet.cell(row=row, column=1).value = experiment.name
        worksheet.cell(row=row, column=2).value = experiment.time
        worksheet.cell(row=row, column=3).value = experiment.film_thickness
        worksheet.cell(row=row, column=4).value = experiment.film_area
        for i, col in enumerate(range(5, 23, 2)):
            worksheet.cell(row=row, column=col).value, worksheet.cell(row=row, column=col+1).value = \
                experiment.values[exp_vals[i]]


def get_efficiencies(experiments, worksheet):
    col_titles = ['Experiment', 'Time (s)', 'Film Thickness (mm)', 'Film Area (cm2)', 'Isc/PV (%)', 'dIsc/PV(%)',
                  'Voc/PV (%)', 'dVoc/PV (%)', 'Pmax/PV (%)', 'dPmax/PV (%)', 'FF/PV (%)', 'dFF/PV (%)', 'Tavg/PV (%)',
                  'dTavg/PV (%)', 'I1avg/PV (%)', 'dI1avg/PV (%)', 'I2avg/PV (%)', 'dI2avg/PV (%)', 'I3avg/PV (%)',
                  'dI3avg/PV (%)', 'I4avg/PV (%)', 'dI4avg/PV (%)']
    exp_vals = ['Delta I_sc', 'Delta V_oc', 'Delta P_max', 'Delta Fill Factor', 'Delta T_avg', 'Delta I_1_avg',
                'Delta I_2_avg', 'Delta I_3_avg', 'Delta I_4_avg']
    for col, title in enumerate(col_titles, 1):
        worksheet.cell(row=2, column=col).value = title
    for row, experiment in enumerate(experiments.values(), 3):
        if experiment.is_reference:
            worksheet['A1'] = "Reference:"
            worksheet['B1'] = experiment.name
        worksheet.cell(row=row, column=1).value = experiment.name
        worksheet.cell(row=row, column=2).value = experiment.time
        worksheet.cell(row=row, column=3).value = experiment.film_thickness
        worksheet.cell(row=row, column=4).value = experiment.film_area
        for i, col in enumerate(range(5, 23, 2)):
            worksheet.cell(row=row, column=col).value, worksheet.cell(row=row, column=col+1).value = \
                experiment.efficiencies[exp_vals[i]]


def get_data(experiment, worksheet):
    col_titles = ['Trace', 'Time (s)', 'Isc (A)', 'dIsc (A)', 'Voc (V)', 'dVoc (V)', 'Pmax (W)', 'dPmax (W)', 'FF',
                  'dFF', 'Tavg (C)', 'dTavg (C)', 'I1avg (W/m2)', 'dI1avg (W/m2)', 'I2avg (W/m2)', 'dI2avg (W/m2)',
                  'I3avg (W/m2)', 'dI3avg (W/m2)', 'I4avg (W/m2)', 'dI4avg (W/m2)']
    exp_vals = ['Short Circuit Current I_sc (A)', 'Open Circuit Voltage V_oc (V)', 'Maximum Power P_max (W)',
                'Fill Factor', 'Average Temperature T_avg (C)', 'Average Irradiance I_1_avg (W/m2)',
                'Average Irradiance I_2_avg (W/m2)', 'Average Irradiance I_3_avg (W/m2)',
                'Average Irradiance I_4_avg (W/m2)']
    worksheet.merge_cells('A1:H1')
    worksheet['A1'] = os.path.join(experiment.folder_path, experiment.name)
    worksheet['A2'] = 'Film Thickness (mm)'
    worksheet['B2'] = experiment.film_thickness
    worksheet['C2'] = 'Film Area (cm2)'
    worksheet['D2'] = experiment.film_area

    for col, title in enumerate(col_titles, 1):
        worksheet.cell(row=3, column=col).value = title
    row = 4
    for row, trace in enumerate([trace for trace in experiment.traces.values() if trace.is_included], 4):
        worksheet.cell(row=row, column=1).value = trace.name
        worksheet.cell(row=row, column=2).value = trace.time
        for i, col in enumerate(range(3, 21, 2)):
            worksheet.cell(row=row, column=col).value, worksheet.cell(row=row, column=col+1).value = \
                trace.values[exp_vals[i]]
    worksheet.cell(row=row+1, column=1).value = 'Average'
    worksheet.cell(row=row+1, column=2).value = experiment.time
    for i, col in enumerate(range(3, 21, 2)):
        worksheet.cell(row=row+1, column=col).value, worksheet.cell(row=row+1, column=col+1).value = \
            experiment.values[exp_vals[i]]

    worksheet.merge_cells('AA1:AB1')
    worksheet['AA1'] = 'Average IV'
    worksheet['AA2'] = 'Voltage (V)'
    worksheet['AB2'] = 'Current (A)'
    for i in range(1, len(experiment.average_data.index)+1):
        worksheet.cell(row=i+3, column=27).value = experiment.average_data['Voltage (V)'][i]
        worksheet.cell(row=i+3, column=28).value = experiment.average_data['Current (A)'][i]

